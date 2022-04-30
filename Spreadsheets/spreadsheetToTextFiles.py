#! python3
# Open and read a spreadsheet
# The contents of column 1 and roww should be written into an individual text file

import openpyxl
from pathlib import Path

# create workbook
wb = openpyxl.Workbook()
sheet = wb.active

listOfTextFiles = []

# Create a workbook 5x5 with dummy text
for row in range(1, 6): 
	for col in range(1, 6):
		file = sheet.cell(row=row, column=col).value = f'text row:{row}, col:{col}'
		
		listOfTextFiles.append(file)
wb.save('testSS.xlsx')

for x in range(sheet.min_row, sheet.max_row + 1):
	textFile = open(f'ssToTextFile{x-1}.txt', 'w')
	for y in list(sheet.columns)[x-1]:
		textFile.write(str(y.value)+ '\n')
		print(y.value)

textFile.close()