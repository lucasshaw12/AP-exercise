#! python3
import openpyxl

################################################
# Getting sheets from the workbook
# wb = openpyxl.load_workbook('example.xlsx')
# print(wb.sheetnames)

# sheet = wb['Sheet3']
# print(sheet)
# print(type(sheet))
# print(sheet.title)

# anotherSheet = wb.active # Sheet that is first when the file is opened
# print(anotherSheet)

################################################
# Getting cells from sheets
# wb = openpyxl.load_workbook('example.xlsx')
# sheet = wb['Sheet1']
# print(sheet['A1'])
# print(sheet['A1'].value)

# print()
# c = sheet['B1']
# print(c.value)
# print('Row %s, Column %s is %s' % (c.row, c.column, c.value))
# print('Cell %s is %s' % (c.coordinate, c.value))
# print(sheet['C1'].value)

# print()
# print(sheet.cell(row=1, column=2))
# print(sheet.cell(row=1, column=2).value)
# for i in range(1, 8, 2):
# 	print(i, sheet.cell(row=i, column=2).value)

# print(sheet.max_row)
# print(sheet.max_column)

################################################
# Converting between column letters and numbers
# from openpyxl.utils import get_column_letter, column_index_from_string
# print(get_column_letter(1))
# print(get_column_letter(2))
# print(get_column_letter(27))
# print(get_column_letter(900))
# print()

# wb = openpyxl.load_workbook('example.xlsx')
# sheet = wb['Sheet1']
# print(get_column_letter(sheet.max_column))
# print(column_index_from_string('A'))
# print(column_index_from_string('AA'))

################################################
# Getting rows and columns from the sheet
# wb = openpyxl.load_workbook('example.xlsx')
# sheet = wb['Sheet1']
# print(tuple(sheet['A1':'C3']))
# print()
# for rowOfCellObjects in sheet['A1':'C3']:
# 	for cellObj in rowOfCellObjects:
# 		print(cellObj.coordinate, cellObj.value)
# 	print('--- END OF ROW ---')


# sheet = wb.active
# print(list(sheet.columns)[1])
# for cellObj in list(sheet.columns)[1]:
# 	print(cellObj.value)


################################################
# Creating and saving excel documents
# wb = openpyxl.Workbook()

# print(wb.sheetnames) # It starts with one sheet named 'sheet'
# sheet = wb.active
# print(sheet.title) 
# sheet.title = 'Spam Bacon Eggs Sheet'
# print(wb.sheetnames)

# sheet = wb.active
# sheet.title = 'Spam Spam Spam'
# wb.save('example_copy.xlsx') # Save the workbook


################################################
# Creating and removing sheets
# wb = openpyxl.Workbook()
# print(wb.sheetnames)
# wb.create_sheet() # Add a new sheet
# print(wb.sheetnames)

# wb.create_sheet(index=0, title='First Sheet') # Create a new sheet at index 0 with the name 'First sheet'
# print(wb.sheetnames)
# wb.create_sheet(index=2, title='Middle Sheet')
# print(wb.sheetnames)

# del wb['Middle Sheet']
# del wb['Sheet1']
# print(wb.sheetnames)

################################################
# Writing values to cells
# wb = openpyxl.Workbook()
# sheet = wb['Sheet']
# sheet['A1'] = 'Hello World!' # Edit the cells value
# print(sheet['A1'].value) # Print the cells value

################################################
# Setting font styles of cells
# from openpyxl.styles import Font
# wb = openpyxl.Workbook()
# sheet = wb['Sheet']
# italic24Font = Font(size=24, italic=True) # Create a Font
# sheet['A1'].font = italic24Font # Apply the font to cell A1
# sheet['A1'] = 'Hello world!'
# wb.save('styles.xlsx')

################################################
# Font objects
# from openpyxl.styles import Font
# wb = openpyxl.Workbook()
# sheet = wb['Sheet']

# fontObj1 = Font(name='Times New Roman', bold=True)
# sheet['A1'].font = fontObj1
# sheet['A1'] = 'Bold Times New Roman'

# fontObj2 = Font(size=24, italic=True)
# sheet['B3'].font = fontObj2
# sheet['B3'] = '24 pt Italic'

# wb.save('styles.xlsx')

################################################
# Formulas
# wb = openpyxl.Workbook()
# sheet = wb.active
# sheet['A1'] = 200
# sheet['A2'] = 300
# sheet['A3'] = '=SUM(A1:A2)' # Set the formula
# wb.save('writeFormula.xlsx')

################################################
# Adjusting rows and columns
# wb = openpyxl.Workbook()
# sheet = wb.active
# sheet['A1'] = 'Tall row'
# sheet['B2'] = 'Wide column'
# sheet.row_dimensions[1].height = 70
# sheet.column_dimensions['B'].width = 20
# wb.save('dimensions.xlsx')

################################################
# Merging and unmerging cells
# wb = openpyxl.Workbook()
# sheet = wb.active
# sheet.merge_cells('A1:D3') # Merge all of these cells
# sheet['A1'] = '12 cells merged together'
# sheet.merge_cells('C5:D5') # Merge these 2 cells together
# sheet['C5'] = 'Two merged cells'
# wb.save('merged.xlsx')

# # unmerge the above merged cells 
# sheet.unmerge_cells('A1:D3')
# sheet.unmerge_cells('C5:D5')
# wb.save('merged.xlsx')

################################################
# Freezing panes
# wb = openpyxl.load_workbook('produceSales.xlsx')
# sheet = wb.active
# sheet.freeze_panes = 'A2'
# wb.save('freezeExample.xlsx')

################################################
# Charts
# wb = openpyxl.Workbook()
# sheet = wb.active
# for i in range(1, 11): # Create some info in column A
# 	sheet['A' + str(i)] = i

# refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=10) # Create the reference object

# seriesObj = openpyxl.chart.Series(refObj, title='First series') # Create the series object

# chartObj = openpyxl.chart.BarChart() # Create the chart object
# chartObj.title = 'My Chart'
# chartObj.append(seriesObj)

# sheet.add_chart(chartObj, 'C5')
# wb.save('sampleChart.xlsx')









